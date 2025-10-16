import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

# ------------------------ COLUNAS ------------------
colunas = ['ROTAS', 'MOTORISTA', 'PLACA', 'CÓDIGO GLAD', 'REALIZADO']

# --------------------- DADOS POR COLUNA -----------------
rotas = ['ROTA 01', 'ROTA 02', 'ROTA 03', 'ROTA 04', 'ROTA 05', 'ROTA 06' ,'ROTA 07', 'ROTA 08', 'ROTA 09', 'ROTA 10', 
         'ROTA 11', 'ROTA 12', 'ROTA 13', 'ROTA 14', 'ROTA 15', 'ROTA 16', 'ROTA 17', 'ROTA 18', 'ROTA 19', 'ROTA 20', 
         'ROTA 21', 'ROTA 22', 'ROTA 23', 'ROTA 24', 'ROTA 25', 'ROTA 26', 'ROTA 27', 'ROTA 28', 'ROTA 29', 'ROTA 30']

motoristas = ['GILMAR', 'GIOVANE', 'MARCEL', 'FELIPE', 'FELIPE', 'LUIZ', 'ELTON', 'JOEL', 'FABIANO COTRIM', 'WAGNER',
              'ADRIANO','FERMINO', 'ISMAEL', 'JULIO', 'GILSON', 'FRANCISCO', 'LEANDRO VIDAL', 'VANNER', 'ERICARLOS', 'MARCELO',
              'GABRIEL', 'LIRA', 'VANDERLEY', 'ABNER', 'CARLOS', 'RENATO', 'PAULO', 'RAFAEL', 'DIEGO', 'ANDRE']

placas = ['RML5C29', 'TKW-3F68', 'TJV6E67', 'ETU0E41', 'EJI9D42', 'SVH-1J39', 'SEJ-6G42', 'TJC5E04', 'TJE7I95', 'RYK-8A31',
          'TMD7J17', 'TJT7B89', 'SVF5G74', 'TMC0I17','SWX-2H43', 'TMD7J17', 'EVW0996', 'SHN-2F65', 'TJG3H35', 'TJU-3I89',
          'STX-8D35', 'SVL-6G42', 'ABC-1234', 'DEF-5678', 'GHI-9012', 'JKL-3456', 'MNO-7890', 'PQR-2345', 'STU-6789', 'VWX-0123']

cod_glad = ['GLAD001', 'GLAD002', 'GLAD003', 'GLAD004', 'GLAD005', 'GLAD006', 'GLAD007', 'GLAD008', 'GLAD009', 'GLAD010',
            'GLAD011', 'GLAD012', 'GLAD013', 'GLAD014', 'GLAD015', 'GLAD016', 'GLAD017', 'GLAD018', 'GLAD019', 'GLAD020',
            'GLAD021', 'GLAD022', 'GLAD023', 'GLAD024', 'GLAD025', 'GLAD026', 'GLAD027', 'GLAD028', 'GLAD029', 'GLAD030']

realizado = ['Sim', 'Não', 'Sim', 'Não', 'Sim', 'Não', 'Sim', 'Não', 'Sim', 'Não',
             'Sim', 'Não', 'Sim', 'Não', 'Sim', 'Não', 'Sim', 'Não', 'Sim', 'Não',
             'Sim', 'Não', 'Sim', 'Não', 'Sim', 'Não', 'Sim', 'Não', 'Sim', 'Não']



# ---------- TRANSFORMAR LISTAS DE COLUNAS EM LISTA DE LINHAS ----------
linhas = list(zip(rotas, motoristas, placas, cod_glad, realizado))

# ---------- CRIAR DATAFRAME ----------
df = pd.DataFrame(linhas, columns=colunas)

# ---------- SALVAR EXCEL ----------
arquivo_excel = 'REPORT_GLAD.xlsx'
df.to_excel(arquivo_excel, index=False, startcol=1, startrow=4)

# ---------- FORMATAR CABEÇALHO (NEGRITO + BORDAS ESPESSAS) ----------
wb = load_workbook(arquivo_excel)
ws = wb.active

negrito = Font(bold=True)
borda_espessa = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)

#----------- FORMATAR CÉLULAS (TODAS AS BORDAS) --------------
thin_border = Border (
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

#----------- APLICAR A UM INTERVALO DE CÉLULAS ---------------
for row in ws ['B6':'F35']:
    for cell in row:
        cell.border = thin_border

celulas_border = ['A4', 'B4', 'C4', 'D4', 'E4', 'F4', 'C1', 'C2', 'C3', 'A5']
for row in celulas_border:
    ws[row].border = thin_border

#----------- PREENCHIMENTOS DE FUNDO --------------------------
verde = PatternFill(start_color='ABE7B1', end_color='ABE7B1', fill_type='solid')
vermelho = PatternFill(start_color='FFBDBD', end_color='FFBDBD', fill_type='solid')
cinza = PatternFill(start_color='EDEDED', end_color='EDEDED', fill_type='solid')
amarelo = PatternFill(start_color='FFFF57', end_color='FFFF57', fill_type='solid')

#----------- FORMATAÇÃO CONDICIONAL PARA O FUNDO --------------
for row in ws['A1':'F40']:
    for cell in row:
        if cell.value == 'Sim':
            cell.fill = verde
        elif cell.value == 'Não':
            cell.fill = vermelho


#----------- FORMATAÇÃO CONDICIONAL PARA AS COLUNAS --------------
celulas_cinza = ['B1', 'B2', 'B3', 'B4', 'C1', 'C2', 'C3', 'A5', 'B5', 'C5', 'D5', 'E5', 'F5', 'A4', 'C4', 'D4', 'F4', 'E4']
for ref in celulas_cinza:
    ws[ref].fill = cinza

# Percorrer as linhas de 5 a 40
for linha in range(5, 41):  # 41 porque o range não inclui o último número
    valor_F = ws[f'F{linha}'].value  # pegar valor da coluna F
    if valor_F is not None:
        # Normaliza o texto para evitar problemas com espaços ou minúsculas
        if str(valor_F).strip().upper() == "NÃO":
            ws[f'C{linha}'].fill = amarelo  # pintar a célula da coluna C


#----------- CÉLULAS SEPARADAS -------------
valores = [
    (1, 2, 'TOTAL ROTAS'),
    (5, 1, '=HOJE()'),
    (2, 2, 'TOTAL ROTAS ABERTAS'),
    (3, 2, 'MÉDIA ROTAS ABERTAS')
]

for row, col, val in valores:
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(bold=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


#---------------- CENTRALIZAR TODAS AS CÉLULAS ------------------
centralizado = Alignment(horizontal='center', vertical='center')

for row in ws.iter_rows():  # iter_rows() percorre todas as linhas existentes
    for cell in row:
        cell.alignment = centralizado

#---------ESPAÇAMENTO COLUNAS (percorre todas as colunas da planilha) --------------
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)  # pega a letra da coluna
    
    for cell in col:
        if cell.value:
            # converte para string e mede o tamanho
            max_length = max(max_length, len(str(cell.value)))
    
    # define a largura da coluna com um pouco de folga (+2)
    ws.column_dimensions[col_letter].width = max_length + 2


#------------- LARGURA DA LINHA 4 -----------------
ws.row_dimensions[4].height = 4

#----------------- RETIRAR LINHAS DE GRADE ---------------
ws.sheet_view.showGridLines = False

#----------------------------------------------- FÓRMULAS ----------------------------------------------
ws['C3'].number_format = numbers.FORMAT_PERCENTAGE_00  # 2 casas decimais
#------------------------------------------- TOTAL DE ROTAS --------------------------------------------
ws['C1'] = '=COUNTA(D6:D40)'
#---------------------------------------- TOTAL ROTAS ABERTAS ------------------------------------------
ws['C2'] = '=COUNTIF(F6:F40, "Sim")'
#---------------------------------------- MÉDIA ROTAS ABERTAS ------------------------------------------
ws['C3'] = '=IF(C1=0, 0, C2/C1)'

# ---------- SALVAR PLANILHA FINAL ----------
wb.save(arquivo_excel)
print(f"✅ Planilha '{arquivo_excel}' criada!")
